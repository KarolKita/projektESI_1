# biblioteka do wczytywania danych
from openpyxl import load_workbook
# biblioteka do liczenia logarytmu
import math
# biblioteka sluzaca do graficznego przdstawienia drzewa
import tkinter as tk


# klasa stworzona do przechowywania oraz tworzenia drzewa decyzyjnego
class Drzewko:
    def __init__(self, korzen):
        self.korzen = korzen  # atrybut z najwieksza laczna entropia
        self.przeslanka = None  # przeslanka dla atrybutu
        self.tak = None  # "galaz tak" przechowuje atrybut/konkluzje
        self.nie = None  # "galaz nie" przechowuje atrybut/konkluzje
        self.tab_tak = None  # pole przechowuje tablice po podziale
        self.tab_nie = None  # pole przechowuje tablice po podziale


# ladowanie danych z excela
wb = load_workbook('../dane_samochody.xlsx')
arkusz = wb.active

# 'tablica' z danymi
tablica = {}

# zmienne pomocnicze
ilo_wierszy = 21  # ilosc wierszy w arkuszu
ilo_kolumn = 164  # ilosc kolumn w arkuszu
przesl = arkusz.cell(1, 1).value  # pierwsza przeslanka w arkuszu
szukana = 'Segment'  # szukana przeslanka

# wczytanie danych do tablicy wiersz po wierszu z arkusza
# gdzie w pierwszej kolumnie sa przeslanki
# w drugiej kolumnie sa atrybuty
# a w kolumnach od 3 do 164 sa przypadki
for nr_wiersza in range(1, ilo_wierszy+1):
    # wczytanie przeslanki
    if arkusz.cell(nr_wiersza, 1).value is not None:
        przesl = arkusz.cell(nr_wiersza, 1).value
        tablica[przesl] = {}  # zagniezdzanie slownika do przechowywania atrybutow

    # wczytanie atrybutu
    tablica[przesl][arkusz.cell(nr_wiersza, 2).value] = []

    # wczytanie przypadkow
    for wiersz in arkusz.iter_rows(min_row=nr_wiersza, min_col=3, max_row=nr_wiersza, max_col=ilo_kolumn):
        for przypadek in wiersz:
            tablica[przesl][arkusz.cell(nr_wiersza, 2).value].append(przypadek.value)


# wypisanie tablicy
def wypisz(tab):
    for p in tab:
        print(p, ": ")
        for a in tab[p]:
            print("     ", a, ": ", end=" ")
            print(tab[p][a])


# liczenie entropi 'I' dla warunku j
# tab - 'tablica' z danymi
# szuk - szukana przeslanka
def entropia(tab, szuk):
    tab_ni = {}  # slownik z ilosciami przypadkow dla danego atrybutu
    n = len(tab['Segment']["A"])  # ilosc kolumn z przypadkami
    # zliczanie ilosci przypadkow dla atrybutu
    for atr in tab[szuk]:
        tab_ni[atr] = sum(tab[szuk][atr])
    # liczenie entropi 'I'
    entr = 0
    for atr in tab_ni:
        if tab_ni[atr] > 0:
            entr -= (tab_ni[atr] / n) * math.log(tab_ni[atr]/n, 2)

    return entr  # zwracana wartosc entropi dla warunku j


# liczenie entropi potwierdzajacych i zaprzeczajacych warunek j
# tab - 'tablica' z danymi
# szuk - szukana przeslanka
# przes - przeslanka
# szuk_atr - atrybut dla ktorego liczona jest entropia
def entr_potw_zaprz(tab, szuk, przes, szuk_atr):
    n = len(tab[przes][szuk_atr])  # liczba przypadkow
    nr = 0  # numer kolumny sprawdzanego przypadku
    # tworzenie slownika do przechowywania ilosci przypadkow potwierdzajacych i zaprzeczajacych warunek j
    tab_ni = {}
    for atr in tab[szuk]:
        tab_ni[atr] = {"n+": 0, "n-": 0}

    # zliczanie ilosci przypadkow dla przekazanego atrybutu
    # w tej petli bedziemy chodzic po przypadkach atrybutu
    # gdzie p_w to przypadek w nr-kolumnie
    for p_w in tab[przes][szuk_atr]:
        # w tej sprawdzamy przypadki dla atrybutow z konkluzji w nr-kolumnie
        for atr in tab[szuk]:
            # szukamy atrybutow dla ktorych wartosc przypadku jest rowna '1'
            if tab[szuk][atr][nr] == 1:
                if p_w == 1:
                    # zwiekszamy ilosc elementow potwierdzajacych warunek
                    tab_ni[atr]['n+'] += 1
                else:
                    # zwiekszamy ilosc elementow zaprzeczających warunek
                    tab_ni[atr]['n-'] += 1

        nr += 1  # zwiekszamy numer kolumny

    # liczenie entropi potwierdzajacych oraz zaprzeczajacych warunek
    n_potw = sum(tab[przes][szuk_atr])  # laczna ilosc elementow potwierdzajacych warunek
    n_zaprz = n - n_potw  # laczna ilosc elementow zaprzeczajacych warunek
    entr_potw = 0  # entropia po potwierdzniu warunku
    entr_zaprz = 0  # entropia po zaprzeczeniu warunku

    for atr in tab[szuk]:
        # liczenie entropi po potwierdzeniu warunku
        if tab_ni[atr]['n+'] > 0:
            entr_potw -= (tab_ni[atr]['n+'] / n_potw) * math.log((tab_ni[atr]['n+'] / n_potw), 2)
        # liczenie entropi po zaprzeczeniu warunku
        if tab_ni[atr]['n-'] > 0:
            entr_zaprz -= (tab_ni[atr]['n-'] / n_zaprz) * math.log((tab_ni[atr]['n-'] / n_zaprz), 2)

    # zwrocenie sumy entropi potwierdzajacych oraz zaprzeczajacych warunek j
    return (n_potw / n) * entr_potw + (n_zaprz / n) * entr_zaprz


# funkcja zwracajaca atrybut i przeslanke z nawieksza laczna entropia
# tab - 'tablica' z danymi
# szuk - szukana przeslanka
# p - dodatkowy parametr ktory zawiera przeslanke "rodzica"
def max_laczna_entropia(tab, szuk, p=None):
    entr_maks = -1  # najwieksza laczna entropia
    atr_maks = ''  # atrybut z najwieksza entropia
    przes_maks = ''  # przeslanka dla ktorej atrybut ma najwieksza laczna entropie

    # wartosc entropi
    entr = entropia(tab, szuk)

    # liczenie oraz szukanie najwiekszej lacznej wartosci entropi
    for przes in tab:
        # omijanie konkluzji oraz unikanie powtorzen atrybutow z jednej przeslanki w jednym wezle
        if przes != szuk and przes != p:
            for atr in tab[przes]:
                # szukanie atrybutu z najwieksza wartoscia lacznej entropi
                if entr_maks < (entr - entr_potw_zaprz(tab, szuk, przes, atr)):
                    entr_maks = entr - entr_potw_zaprz(tab, szuk, przes, atr)
                    atr_maks = atr
                    przes_maks = przes

    # zwrocenie atrybutu oraz przeslanki z najwieksza laczna entropia
    return atr_maks, przes_maks


# funkcja dzielaca tabele
# tab 'tablica' z danymi
# przes_p przeslanka dla ktorej atrybut posiada najwieksza laczna entropie
# atr_p atrybut na podstawie ktorego dzielimy tabele
def podzial_tab(tab, przes_p, atr_p):
    tab_tak = {}  # tablica z danymi potwierdzajacymi warunek
    tab_nie = {}  # tablica z danymi zaprzeczajacymi warunek

    # wczytywanie przeslanek i atrybutow do tabel
    for przes in tab:
        tab_tak[przes] = {}
        tab_nie[przes] = {}
        for atr in tab[przes]:
            tab_tak[przes][atr] = []
            tab_nie[przes][atr] = []

    nr_kol = 0  # numer kolumny

    # wczytanie przypadkow do tablic
    for przyp_p in tab[przes_p][atr_p]:
        for przes in tab:
            for atr in tab[przes]:
                if przyp_p == 1:
                    # tablica z elementami potwierdzajacymi
                    tab_tak[przes][atr].append(tab[przes][atr][nr_kol])
                else:
                    # tablica z elementami zaprzeczajacymi
                    tab_nie[przes][atr].append(tab[przes][atr][nr_kol])
        nr_kol += 1  # zwiekszanie numeru kolumny

    # zwrocenie tablicy z potwierdzajacych i zaprzeeczajacych warunek
    return tab_tak, tab_nie


# funkcja sprawdzajaca czy jedna z szukanych konkluzji posiada same '1'
# tab - 'tablica' z danymi
# szuk - szukana przeslanka
def sprawdz(tab, szuk):
    for atr in tab[szuk]:
        if len(tab[szuk][atr]) == sum(tab[szuk][atr]):
            return 1


# funkcja zwraca konkluzje z samymi '1' w przypadkach
# tab - 'tablica' z danymi
# szuk - szukana przeslanka
def zwroc_konkluzje(tab, szuk):
    for konluzja in tab[szuk]:
        if len(tab[szuk][konluzja]) == sum(tab[szuk][konluzja]):
            return konluzja


# tworzenie korzenia
atr_korzen, przes_korz = max_laczna_entropia(tablica, szukana)
drzewo = Drzewko(atr_korzen)
drzewo.przeslanka = przes_korz


# funkcja tworzaca drzewo
# korz - to obiekt klasy Drzewko
# tab - 'tablica' z danymi
# szuk - szukana przeslanka
def tworz_drzewo(korz, tab, szuk):
    # podzial tabeli 'rodzica'
    korz.tab_tak, korz.tab_nie = podzial_tab(tab, korz.przeslanka, korz.korzen)

    # tworzenie kolejnych galezi drzewa
    # galezie 'tak'
    # sprawdzanie czy w otrzymanej tablicy wystepuje jedna wartosc konkluzji
    if sprawdz(korz.tab_tak, szuk) == 1:
        # dodawanie konkluzji do drzewa
        korz.tak = zwroc_konkluzje(korz.tab_tak, szuk)
    else:
        # szukanie atrybutu z najwieksza laczna entropia
        temp_t = max_laczna_entropia(korz.tab_tak, szuk)
        # sprawdzenie czy aktualna przeslanka jest rozna od przeslanki rodzica
        if temp_t[1] == korz.przeslanka:
            temp_t = max_laczna_entropia(korz.tab_tak, szuk, temp_t[1])
        # tworzenie nowego obiektu Drzewka
        korz.tak = Drzewko(temp_t[0])
        korz.tak.przeslanka = temp_t[1]
        # wywolanie rekurencyjne funkcji tworz_drzewo()
        tworz_drzewo(korz.tak, korz.tab_tak, szuk)

    # galezie 'nie'
    # tworzona analogicznie jak "galaz tak"
    if sprawdz(korz.tab_nie, szuk) == 1:
        korz.nie = zwroc_konkluzje(korz.tab_nie, szuk)
    else:
        temp_n = max_laczna_entropia(korz.tab_nie, szuk)
        if temp_n[1] == korz.przeslanka:
            temp_n = max_laczna_entropia(korz.tab_tak, szuk, temp_n[1])
        korz.nie = Drzewko(temp_n[0])
        korz.nie.przeslanka = temp_n[1]
        tworz_drzewo(korz.nie, korz.tab_nie, szuk)

    # zwrocenie obiektu klasy Drzewko
    return korz


# tworzenie drzewa
tworz_drzewo(drzewo, tablica, szukana)

# lista zawierajaca dane do rysowania drzewa
dane = []


# tworzenie danych potrzebnych do graficznego przedstawienia drzewa,
# gdzie korz to obiekt Drzewa
# d - lista z danymi
# x - wspolrzedna x
# y - wspolrzedna y
# wart - okresla galaz 'tak' lub 'nie' do rodzica
# funkcja zwraca liste zawierajaca listy z nazwami etykiet oraz wspolrzednymi 'x' i 'y'
def dane_do_rysowania(korz, d, x, y, wart=None):
    # dane korzenia
    d.append([korz.korzen, x, y, wart])

    # dane wezla 'tak'
    if isinstance(korz.tak, Drzewko):
        dane_do_rysowania(korz.tak, d, x - 300, y + 120, 1)
    else:
        d.append([korz.tak, x - 100, y + 75, 1])

    # dane wezla 'nie'
    if isinstance(korz.nie, Drzewko):
        dane_do_rysowania(korz.nie, d, x + 300, y + 120, 0)
    else:
        d.append([korz.nie, x + 100, y + 75, 0])

    # zwrocenie listy danych
    return d


# slownik zawierajacy nazwy etykiet
nazwy = {"15 <": "dystans < 15 km",
         "15 - 35": "dystans miedzy 15 a 35 km",
         "35 >=": "dystans >= 35 km",
         " 1": "liczba przewozonych osob = 1",
         " 2": "liczba przewozonych osob = 2",
         "3 >=": "liczba przewozonych osob >= 3",
         "50 <": "wielkosc bagazu < 50 kg",
         "50 - 100": "wielkosc bagazu miedzy 50 a 100 kg",
         "100 >=": "wielkosc bagazu >= 100 kg",
         "<= 50": "srednia predkosc <= 50 km/h",
         "50 - 70": "srednia predkosc miedzy 50 a 70 km/h",
         "70 >=": "srednia predkosc >= 70 km/h",
         "ekspresowa": "droga ekspresowa",
         "główna": "droga glowna",
         "lokalna": "droga lokalna",
         "gruntowa": "droga gruntowa",
         "A": "auto małe",
         "B": "auto miejskie",
         "C": "auto kompaktowe",
         "D": "auto rodzinne",
         "J": "auto terenowe"}

# wywołanie funkcji
dane_do_rysowania(drzewo, dane, 425, 0)

# graficzne przedstawienie drzewa
# tworzenie okna tkinter
okno = tk.Tk()

# dodanie tytulu i rozmiarow okna
okno.title("Drzewko decyzyjne - dobór typu samochodu do potrzeb")
okno.geometry("1600x800")
okno.resizable(False, False)

# pakiet uzywany do rysowania lini
canvas = tk.Canvas(okno, width=1600, height=800)
canvas.pack()

# tworzenie graficzne drzewka
for etykieta in dane:
    # tworzenie etykiet
    tk.Label(master=okno,
             text=nazwy[etykieta[0]],
             font=20,
             padx=10,
             pady=5,
             borderwidth=2,
             relief="solid").place(x=etykieta[1], y=etykieta[2])

    # wspolrzedne startowe
    x_start = etykieta[1] + 50
    y_start = etykieta[2] + 20
    # wspolrzedne koncowe
    x_stop = x_start
    y_stop = y_start

    # tworzenie lini oraz etykiet 'tak' i 'nie'
    if etykieta[0] == drzewo.korzen:
        continue
    elif etykieta[0] == 'A' or etykieta[0] == 'B' or etykieta[0] == 'C' or etykieta[0] == 'D' or etykieta[0] == 'J':
        y_stop -= 75
        y_etyk = 50
        if etykieta[3] == 1:
            x_stop += 100
            kolor = "green"
            tekst = "TAK"
            x_etyk = 35
        else:
            x_stop -= 100
            kolor = "red"
            tekst = "NIE"
            x_etyk = -65
    else:
        y_stop -= 120
        y_etyk = 75
        if etykieta[3] == 1:
            x_stop += 300
            kolor = "green"
            tekst = "TAK"
            x_etyk = 150
        else:
            x_stop -= 300
            kolor = "red"
            tekst = "NIE"
            x_etyk = -170

    canvas.create_line(x_start, y_start, x_stop, y_stop, fill=kolor, width=3)
    tk.Label(master=okno, text=tekst, font=5).place(x=x_start + x_etyk, y=y_start - y_etyk)

okno.mainloop()
